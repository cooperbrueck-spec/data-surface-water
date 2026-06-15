"""Build an old-style SWP Bulletin 132 workbook from B132-25 Appendix B.

The Stata SWP cleaner expects the historical B132 workbook shape:
Contractors, B4, and B-5B sheets, with B4/B-5B imported from A3:AN79.
This script preserves that shape while sourcing B4/B-5B values from the
B132-25 PDF. B132-25 labels the former Castaic Lake column as Santa Clarita;
the workbook keeps Castaic Lake for downstream compatibility.
"""

from __future__ import annotations

import re
from copy import copy
from itertools import combinations
from pathlib import Path

import openpyxl
import pdfplumber
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[2]
SWP_DIR = ROOT / "raw" / "swp"
TEMPLATE = SWP_DIR / "bulletin_132" / "B132-21 Tables.xlsx"
SOURCE_PDF = SWP_DIR / "bulletin_132" / "B132-25 Appendix B 032626a.pdf"
OUTPUT = SWP_DIR / "bulletin_132" / "B132-25 Tables.xlsx"

YEAR_RE = re.compile(r"^(?:19|20)\d{2}$")
NUMBER_RE = re.compile(r"^-?\d[\d,]*$")

HEADERS = [
    "Calendar Year",
    "Napa",
    "Solano",
    "Total",
    "Alameda- Zone 7",
    "Alameda County",
    "Santa Clara",
    "Total",
    "San Luis Obispo",
    "Santa Barbara",
    "Total",
    "Dudley Ridge",
    "Empire",
    "Kern: Municipal and Industrial",
    "Kern: Agricultural",
    "Kern: Total",
    "Kings",
    "Oak Flat",
    "Tulare",
    "Total",
    "AVEK",
    "Castaic Lake",
    "Coachella",
    "Crestline",
    "Desert",
    "Littlerock",
    "Mojave",
    "Palmdale",
    "San Bernardino",
    "San Gabriel",
    "San Gorgonio",
    "Metropolitan",
    "Ventura",
    "Total",
    "Yuba City",
    "Butte",
    "Plumas",
    "Total",
    "South Bay Area Future Contractor",
    "Grand Total",
]

PAGE_COLUMN_MAP = {
    48: list(range(1, 11)),
    49: list(range(11, 20)),
    50: list(range(20, 30)),
    51: list(range(30, 40)),
    80: list(range(1, 11)),
    81: [11, 12, 13, 14, 16, 17, 18, 19],
    82: list(range(20, 30)),
    83: list(range(30, 40)),
}

TABLE_PAGES = {
    "B4": [48, 49, 50, 51],
    "B-5B": [80, 81, 82, 83],
}


def parse_int(value: str) -> int:
    return int(value.replace(",", ""))


def extract_rows(pdf: pdfplumber.PDF, page_number: int) -> dict[int | str, list[int]]:
    """Extract numeric rows keyed by year or TOTAL from a one-based PDF page."""
    page = pdf.pages[page_number - 1]
    words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False)
    lines: list[list[dict]] = []
    for word in sorted(words, key=lambda item: (item["top"], item["x0"])):
        if not lines or abs(word["top"] - lines[-1][0]["top"]) > 2:
            lines.append([word])
        else:
            lines[-1].append(word)

    rows: dict[int | str, list[int]] = {}
    for line_words in lines:
        line_words = sorted(line_words, key=lambda item: item["x0"])
        tokens = [item["text"] for item in line_words]
        if not tokens:
            continue

        key: int | str | None = None
        if YEAR_RE.match(tokens[0]):
            key = int(tokens[0])
            tokens = tokens[1:]
        elif tokens[0] == "TOTAL":
            key = "TOTAL"
            tokens = tokens[1:]
        else:
            continue

        numeric = [parse_int(token) for token in tokens if NUMBER_RE.match(token)]
        numeric = normalize_page_values(page_number, numeric)
        if numeric:
            rows[key] = numeric

    return rows


def local_total_checks(page_number: int, by_col: dict[int, int]) -> bool:
    checks: list[tuple[int, int]] = []
    if page_number in [48, 80]:
        checks = [
            (by_col[3], by_col[1] + by_col[2]),
            (by_col[7], by_col[4] + by_col[5] + by_col[6]),
            (by_col[10], by_col[8] + by_col[9]),
        ]
    elif page_number == 49:
        checks = [
            (by_col[15], by_col[13] + by_col[14]),
            (by_col[19], by_col[11] + by_col[12] + by_col[15] + by_col[16] + by_col[17] + by_col[18]),
        ]
    elif page_number == 81:
        checks = [
            (by_col[19], by_col[11] + by_col[12] + by_col[13] + by_col[14] + by_col[16] + by_col[17] + by_col[18]),
        ]
    elif page_number in [51, 83]:
        checks = [
            (by_col[33], by_col[30] + by_col[31] + by_col[32]),
            (by_col[37], by_col[34] + by_col[35] + by_col[36]),
        ]

    return all(observed == expected for observed, expected in checks)


def normalize_page_values(page_number: int, values: list[int]) -> list[int]:
    """Restore zeroes dropped by PDF text extraction, verified by local totals."""
    cols = PAGE_COLUMN_MAP[page_number]
    expected = len(cols)
    if len(values) == expected:
        by_col = dict(zip(cols, values))
        if local_total_checks(page_number, by_col):
            return values

    if len(values) > expected:
        return values

    missing = expected - len(values)
    # Insert only zeroes. This handles PDF text extraction losses without
    # inventing nonzero data.
    for positions in combinations(range(expected + 1), missing):
        candidate = values[:]
        for pos in sorted(positions):
            candidate.insert(pos, 0)
        by_col = dict(zip(cols, candidate))
        if local_total_checks(page_number, by_col):
            return candidate

    return values


def fill_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, rows_by_col: dict[int | str, dict[int, int]]) -> None:
    for excel_row in range(5, 79):
        year = sheet.cell(excel_row, 1).value
        if not isinstance(year, int):
            raise ValueError(f"{sheet.title}: expected year in row {excel_row}, found {year!r}")
        for col_idx in range(1, 40):
            sheet.cell(excel_row, col_idx + 1).value = rows_by_col[year][col_idx]

    for col_idx in range(1, 40):
        sheet.cell(79, col_idx + 1).value = rows_by_col["TOTAL"][col_idx]


def apply_b13225_row_styles(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Mark B132-25 actual, projected, and bill-year rows using the source convention."""
    actual_styles = [copy(sheet.cell(64, col)._style) for col in range(1, 41)]  # 2021
    projected_fill = PatternFill(fill_type="solid", fgColor="FFD1D3D4")

    for excel_row in range(65, 79):  # 2022-2035 year rows
        for col in range(1, 41):
            cell = sheet.cell(excel_row, col)
            cell._style = copy(actual_styles[col - 1])
            font = copy(cell.font)
            font.bold = False
            cell.font = font

    for excel_row in range(68, 79):  # 2025-2035 projected
        for col in range(1, 41):
            sheet.cell(excel_row, col).fill = copy(projected_fill)

    for col in range(1, 41):  # 2026 bill year
        cell = sheet.cell(69, col)
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def validate_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    headers = [sheet.cell(3, col).value for col in range(1, 41)]
    if headers != HEADERS:
        raise ValueError(f"{sheet.title}: header mismatch")

    years = [sheet.cell(row, 1).value for row in range(5, 79)]
    if years != list(range(1962, 2036)):
        raise ValueError(f"{sheet.title}: expected years 1962-2035")

    if sheet.cell(79, 1).value != "TOTAL":
        raise ValueError(f"{sheet.title}: row 79 must be TOTAL")

    for row in range(5, 80):
        values = {col: sheet.cell(row, col + 1).value for col in range(1, 40)}
        checks = [
            (values[3], values[1] + values[2]),
            (values[7], values[4] + values[5] + values[6]),
            (values[10], values[8] + values[9]),
            (values[15], values[13] + values[14]),
            (values[19], values[11] + values[12] + values[15] + values[16] + values[17] + values[18]),
            (
                values[33],
                sum(values[col] for col in range(20, 33)),
            ),
            (values[37], values[34] + values[35] + values[36]),
            (values[39], values[3] + values[7] + values[10] + values[19] + values[33] + values[37] + values[38]),
        ]
        for observed, expected in checks:
            if observed != expected:
                raise ValueError(
                    f"{sheet.title}: row {row} total mismatch, observed {observed}, expected {expected}"
                )


def build_table(pdf: pdfplumber.PDF, table_name: str) -> dict[int | str, dict[int, int]]:
    rows_by_col: dict[int | str, dict[int, int]] = {}
    for page_number in TABLE_PAGES[table_name]:
        page_rows = extract_rows(pdf, page_number)
        cols = PAGE_COLUMN_MAP[page_number]
        for key, values in page_rows.items():
            if len(values) != len(cols):
                raise ValueError(
                    f"{table_name} page {page_number} row {key}: "
                    f"expected {len(cols)} values, found {len(values)}"
                )
            rows_by_col.setdefault(key, {})
            for col_idx, value in zip(cols, values):
                rows_by_col[key][col_idx] = value

    expected_keys = list(range(1962, 2036)) + ["TOTAL"]
    for key in expected_keys:
        if key not in rows_by_col:
            raise ValueError(f"{table_name}: missing row {key}")
        if table_name == "B-5B":
            rows_by_col[key][15] = rows_by_col[key][13] + rows_by_col[key][14]
        missing = [col for col in range(1, 40) if col not in rows_by_col[key]]
        if missing:
            raise ValueError(f"{table_name}: row {key} missing columns {missing}")

    return rows_by_col


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    with pdfplumber.open(SOURCE_PDF) as pdf:
        for table_name in ["B4", "B-5B"]:
            rows_by_col = build_table(pdf, table_name)
            fill_sheet(wb[table_name], rows_by_col)
            apply_b13225_row_styles(wb[table_name])
            validate_sheet(wb[table_name])

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
