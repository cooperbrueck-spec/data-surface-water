"""
Update raw/colorado/accounting_reports.xlsx from the Colorado River annual
accounting PDFs.

The Stata pipeline reads the Diversion sheet as annual acre-feet by user.  This
builder preserves the existing workbook rows and inserts new year columns from
Table 5, "State of California - Records of Diversion, Returns, and Consumptive
Use."  Existing historical columns are not recalculated.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl
import pdfplumber


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "raw" / "colorado" / "accounting_reports.xlsx"
PDF_DIR = ROOT / "raw" / "colorado" / "accounting reports"
VALIDATION_CSV = ROOT / "temp" / "colorado_accounting_reports_validation.csv"

UPDATE_YEARS = [2022, 2023, 2024, 2025]
CALIBRATION_YEAR = 2021
TABLE_5_PAGES = {
    # One-based PDF page numbers.  These are the Table 5 California pages,
    # excluding the table-of-contents references.
    2021: [26, 27, 28],
    2022: [27, 28, 29],
    2023: [28, 29, 30],
    2024: [26, 27, 28],
    2025: [26, 27, 28],
}


def clean_line(line: str) -> str:
    return re.sub(r"\s+", " ", line).strip()


def parse_int(token: str) -> int:
    return int(token.replace(",", ""))


def numeric_tokens_after_diversion(line: str) -> list[int]:
    if "Diversion" not in line:
        return []
    tail = line[line.find("Diversion") + len("Diversion") :]
    return [parse_int(x) for x in re.findall(r"-?\d[\d,]*", tail)]


def row_total_from_line(line: str) -> tuple[int, str]:
    values = numeric_tokens_after_diversion(line)
    if not values:
        raise ValueError(f"No numeric diversion values found in row: {line}")

    total = values[-1]
    if len(values) == 13:
        monthly_sum = sum(values[:-1])
        status = "monthly_sum_ok" if monthly_sum == total else f"monthly_sum_mismatch:{monthly_sum}"
    elif len(values) > 13:
        # Some PDF rows split two-digit monthly values into separate digits.
        # The rightmost annual total remains a single token in the source.
        status = f"total_only_token_count:{len(values)}"
    else:
        status = f"total_only_token_count:{len(values)}"
    return total, status


def table_5_lines(pdf_path: Path, year: int) -> list[str]:
    lines: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_number in TABLE_5_PAGES[year]:
            page = pdf.pages[page_number - 1]
            text = page.extract_text(x_tolerance=1, y_tolerance=3, layout=True) or ""
            for line in text.splitlines():
                cleaned = clean_line(line)
                if cleaned:
                    lines.append(cleaned)
    if not lines:
        raise ValueError(f"Could not find Table 5 lines in {pdf_path}")
    return lines


def find_first_index(lines: list[str], needle: str) -> int:
    for idx, line in enumerate(lines):
        if needle in line:
            return idx
    raise ValueError(f"Could not find section heading: {needle}")


def extract_after(
    lines: list[str],
    section: str,
    row_phrase: str,
    occurrence: int = 1,
) -> tuple[int, str, str]:
    start = find_first_index(lines, section)
    seen = 0
    for line in lines[start:]:
        if row_phrase in line and "Diversion" in line:
            seen += 1
            if seen == occurrence:
                total, status = row_total_from_line(line)
                return total, status, line
    raise ValueError(f"Could not find diversion row '{row_phrase}' after '{section}'")


def extract_colorado_diversions(year: int) -> tuple[dict[str, int], list[dict[str, str]]]:
    pdf_path = PDF_DIR / f"{year}.pdf"
    lines = table_5_lines(pdf_path, year)
    checks: list[dict[str, str]] = []

    def add(
        output_name: str,
        section: str,
        row_phrase: str,
        occurrence: int = 1,
    ) -> int:
        total, status, source_line = extract_after(lines, section, row_phrase, occurrence)
        checks.append(
            {
                "year": str(year),
                "user": output_name,
                "value": str(total),
                "check": status,
                "source_line": source_line,
            }
        )
        return total

    values: dict[str, int] = {}
    values["Fort Mojave Indian Reservation"] = add(
        "Fort Mojave Indian Reservation",
        "Fort Mojave Indian Reservation",
        "agriculture use",
    )
    values["City of Needles"] = add("City of Needles", "City of Needles", "Pumped from wells")
    values["Chemehuevi Indian Reservation"] = add(
        "Chemehuevi Indian Reservation",
        "Chemehuevi Indian Reservation",
        "agricultural use",
    )
    values["Metropolitan Water District of Southern California"] = add(
        "Metropolitan Water District of Southern California",
        "The Metropolitan Water District of Southern California",
        "Pumped from Lake Havasu",
    )
    values["Transfer from SDCWA to MWD (originally from IID)"] = 0
    checks.append(
        {
            "year": str(year),
            "user": "Transfer from SDCWA to MWD (originally from IID)",
            "value": "0",
            "check": "row_not_reported_in_table_5",
            "source_line": "",
        }
    )
    values["Bureau of Reclamation and Government Camp"] = add(
        "Bureau of Reclamation and Government Camp",
        "Bureau of Reclamation - Parker Dam and Government Camp",
        "Parker Dam",
    )
    values["Colorado River Indian Reservation"] = add(
        "Colorado River Indian Reservation",
        "Colorado River Indian Reservation",
        "agricultur",
    )
    values["City of Winterhaven"] = add(
        "City of Winterhaven",
        "City of Winterhaven",
        "Pumped from well",
    )
    values["Palo Verde Irrigation District"] = add(
        "Palo Verde Irrigation District",
        "Palo Verde Irrigation District",
        "Palo Verde Dam",
    )
    values["Picacho Development Corporation"] = 0
    checks.append(
        {
            "year": str(year),
            "user": "Picacho Development Corporation",
            "value": "0",
            "check": "row_not_reported_in_table_5",
            "source_line": "",
        }
    )

    if year == 2025:
        values["Yuma Project Reservation Division, Indian Unit"] = add(
            "Yuma Project Reservation Division, Indian Unit",
            "Fort Yuma Indian Reservation (Quechan Indian Tribe)",
            "Imperial Dam",
        )
        values["Yuma Project Reservation Division, Bard Unit"] = add(
            "Yuma Project Reservation Division, Bard Unit",
            "Bard Water District",
            "Imperial Dam",
        )
    else:
        values["Yuma Project Reservation Division, Indian Unit"] = add(
            "Yuma Project Reservation Division, Indian Unit",
            "Yuma Project Reservation Division",
            "Imperial Dam",
            occurrence=1,
        )
        values["Yuma Project Reservation Division, Bard Unit"] = add(
            "Yuma Project Reservation Division, Bard Unit",
            "Yuma Project Reservation Division",
            "Imperial Dam",
            occurrence=2,
        )

    values["Imperial Irrigation District"] = add(
        "Imperial Irrigation District",
        "Imperial Irrigation District",
        "Imperial Dam",
    )
    values["Transfer to San Diego County Water Authority"] = 0
    checks.append(
        {
            "year": str(year),
            "user": "Transfer to San Diego County Water Authority",
            "value": "0",
            "check": "row_not_reported_in_table_5",
            "source_line": "",
        }
    )
    values["Coachella Valley Water District"] = add(
        "Coachella Valley Water District",
        "Coachella Valley Water District",
        "Imperial Dam",
    )

    california_total = add("California Totals", "California Totals", "Diversion")
    named_total = sum(values.values())
    values["Other users"] = california_total - named_total
    checks.append(
        {
            "year": str(year),
            "user": "Other users",
            "value": str(values["Other users"]),
            "check": f"california_total_residual:{california_total}-{named_total}",
            "source_line": "",
        }
    )
    return values, checks


def existing_diversion_values(ws: openpyxl.worksheet.worksheet.Worksheet, year: int) -> dict[str, int]:
    year_col = None
    for col in range(2, ws.max_column + 1):
        if ws.cell(1, col).value == year:
            year_col = col
            break
    if year_col is None:
        raise ValueError(f"Workbook has no Diversion column for {year}")
    return {ws.cell(row, 1).value: ws.cell(row, year_col).value for row in range(2, ws.max_row + 1)}


def ensure_year_columns(ws: openpyxl.worksheet.worksheet.Worksheet, years: list[int]) -> dict[int, int]:
    existing = {ws.cell(1, col).value: col for col in range(2, ws.max_column + 1)}
    missing = [year for year in years if year not in existing]
    if missing:
        ws.insert_cols(2, amount=len(missing))
        for offset, year in enumerate(sorted(missing, reverse=True), start=2):
            ws.cell(1, offset).value = year
    return {ws.cell(1, col).value: col for col in range(2, ws.max_column + 1)}


def write_validation(rows: list[dict[str, str]]) -> None:
    VALIDATION_CSV.parent.mkdir(parents=True, exist_ok=True)
    with VALIDATION_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["year", "user", "value", "check", "source_line"])
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK)
    ws = workbook["Diversion"]
    workbook_users = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    all_checks: list[dict[str, str]] = []
    extracted: dict[int, dict[str, int]] = {}

    calibration_values, calibration_checks = extract_colorado_diversions(CALIBRATION_YEAR)
    all_checks.extend(calibration_checks)
    existing_2021 = existing_diversion_values(ws, CALIBRATION_YEAR)
    mismatches = [
        (user, existing_2021.get(user), calibration_values.get(user))
        for user in workbook_users
        if existing_2021.get(user) != calibration_values.get(user)
    ]
    if mismatches:
        raise ValueError(f"Calibration against 2021 workbook failed: {mismatches}")

    for year in UPDATE_YEARS:
        values, checks = extract_colorado_diversions(year)
        extracted[year] = values
        all_checks.extend(checks)

    year_cols = ensure_year_columns(ws, UPDATE_YEARS)
    for year, values in extracted.items():
        col = year_cols[year]
        for row in range(2, ws.max_row + 1):
            user = ws.cell(row, 1).value
            if user not in values:
                raise ValueError(f"No extracted value for workbook user '{user}' in {year}")
            ws.cell(row, col).value = values[user]

    workbook.save(WORKBOOK)
    write_validation(all_checks)
    print(f"Updated {WORKBOOK}")
    print(f"Wrote validation report to {VALIDATION_CSV}")


if __name__ == "__main__":
    main()
